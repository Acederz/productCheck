"""商品 Excel 导入服务。"""

import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.constants import EXCEL_HEADERS, PLATFORMS, TASK_STATUS_UNASSIGNED
from app.extensions import db
from app.models.task import ClassificationTask, ImportBatch
from app.services.operation_log_service import write_operation_log
from app.utils.excel_helper import cell_str, parse_desc_images, parse_segment

# Excel 列名 -> 模型字段
HEADER_FIELD_MAP = {
    "序号": "row_no",
    "宝贝ID": "product_id",
    "宝贝主图": "main_image",
    "宝贝名称": "product_name",
    "品牌": "brand",
    "是否经营": "is_operating",
    "大类": "category_large",
    "区隔": "category_segment",
    "类别": "category_type",
    "主材质": "material_main",
    "辅材质": "material_aux",
    "包装方式": "packaging",
    "尺寸": "size",
    "卷数": "roll_count",
    "总入数": "total_count",
    "产品属性": "product_attr",
    "宝贝文描图": "desc_images",
    "宝贝链接": "product_url",
    "数据平台": "platform",
}


class ImportService:
    """处理待分类商品 Excel 导入。"""

    def __init__(self, upload_folder: Path):
        self.upload_folder = upload_folder
        self.upload_folder.mkdir(parents=True, exist_ok=True)

    def _gen_batch_no(self) -> str:
        """生成导入批次号。"""
        return datetime.now().strftime("%Y%m%d%H%M%S") + uuid.uuid4().hex[:6]

    def _validate_headers(self, headers: list) -> str | None:
        """校验表头是否与模板一致。"""
        normalized = [cell_str(h) for h in headers[: len(EXCEL_HEADERS)]]
        if len(normalized) < len(EXCEL_HEADERS):
            return f"表头列数不足，需要 {len(EXCEL_HEADERS)} 列"
        for idx, expected in enumerate(EXCEL_HEADERS):
            if normalized[idx] != expected:
                return f"第 {idx + 1} 列表头应为「{expected}」，实际为「{normalized[idx]}」"
        return None

    def _validate_row(self, row_data: dict, excel_row_no: int) -> str | None:
        """行级校验，返回错误信息。"""
        product_id = cell_str(row_data.get("product_id"))
        product_name = cell_str(row_data.get("product_name"))
        platform = cell_str(row_data.get("platform"))

        if not product_id:
            return f"第 {excel_row_no} 行：宝贝ID 不能为空"
        if not product_name:
            return f"第 {excel_row_no} 行：宝贝名称 不能为空"
        if not platform:
            return f"第 {excel_row_no} 行：数据平台 不能为空"
        if platform not in PLATFORMS:
            return f"第 {excel_row_no} 行：数据平台「{platform}」不合法"
        return None

    def _row_to_task(self, row_data: dict, batch_id: int) -> ClassificationTask:
        """将解析后的行数据转为任务对象。"""
        row_no_raw = row_data.get("row_no")
        row_no = None
        if row_no_raw not in (None, ""):
            try:
                row_no = int(float(cell_str(row_no_raw)))
            except ValueError:
                row_no = None

        segment = row_data.get("category_segment")
        if isinstance(segment, str):
            segment = parse_segment(segment)
        elif segment is None:
            segment = []

        return ClassificationTask(
            batch_id=batch_id,
            row_no=row_no,
            product_id=cell_str(row_data.get("product_id")),
            main_image=cell_str(row_data.get("main_image")) or None,
            product_name=cell_str(row_data.get("product_name")),
            brand=cell_str(row_data.get("brand")) or None,
            product_attr=cell_str(row_data.get("product_attr")) or None,
            desc_images=row_data.get("desc_images") or [],
            product_url=cell_str(row_data.get("product_url")) or None,
            platform=cell_str(row_data.get("platform")),
            is_operating=cell_str(row_data.get("is_operating")) or None,
            category_large=cell_str(row_data.get("category_large")) or None,
            category_segment=segment,
            category_type=cell_str(row_data.get("category_type")) or None,
            material_main=cell_str(row_data.get("material_main")) or None,
            material_aux=cell_str(row_data.get("material_aux")) or None,
            packaging=cell_str(row_data.get("packaging")) or None,
            size=cell_str(row_data.get("size")) or None,
            roll_count=cell_str(row_data.get("roll_count")) or None,
            total_count=cell_str(row_data.get("total_count")) or None,
            status=TASK_STATUS_UNASSIGNED,
        )

    def _write_error_report(self, batch_no: str, error_rows: list) -> str | None:
        """生成错误明细 Excel，返回文件路径。"""
        if not error_rows:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "错误明细"
        ws.append(["Excel行号", "错误原因", *EXCEL_HEADERS])

        for item in error_rows:
            ws.append(
                [
                    item["row_no"],
                    item["reason"],
                    *[item["raw"].get(h, "") for h in EXCEL_HEADERS],
                ]
            )

        error_path = self.upload_folder / f"errors_{batch_no}.xlsx"
        wb.save(error_path)
        return str(error_path)

    def import_excel(self, file_storage, user_id: int) -> ImportBatch:
        """执行 Excel 导入。"""
        batch_no = self._gen_batch_no()
        safe_name = file_storage.filename or "import.xlsx"
        saved_path = self.upload_folder / f"{batch_no}_{safe_name}"
        file_storage.save(saved_path)

        batch = ImportBatch(
            batch_no=batch_no,
            file_name=safe_name,
            file_path=str(saved_path),
            status="processing",
            uploaded_by=user_id,
        )
        db.session.add(batch)
        db.session.flush()

        wb = load_workbook(saved_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            batch.status = "failed"
            batch.total_rows = 0
            db.session.commit()
            raise ValueError("Excel 文件为空")

        header_error = self._validate_headers(list(rows[0]))
        if header_error:
            batch.status = "failed"
            db.session.commit()
            raise ValueError(header_error)

        success_tasks = []
        error_rows = []
        total = 0

        for idx, row in enumerate(rows[1:], start=2):
            if row is None or all(c is None or cell_str(c) == "" for c in row):
                continue
            total += 1

            row_data = {}
            raw_map = {}
            for col_idx, header in enumerate(EXCEL_HEADERS):
                val = row[col_idx] if col_idx < len(row) else None
                raw_map[header] = val
                field = HEADER_FIELD_MAP[header]
                if field == "desc_images":
                    row_data[field] = parse_desc_images(val)
                elif field == "category_segment":
                    row_data[field] = parse_segment(val)
                else:
                    row_data[field] = val

            err = self._validate_row(row_data, idx)
            if err:
                error_rows.append({"row_no": idx, "reason": err, "raw": raw_map})
                continue

            success_tasks.append(self._row_to_task(row_data, batch.id))

        if success_tasks:
            db.session.bulk_save_objects(success_tasks)

        error_report = self._write_error_report(batch_no, error_rows)

        batch.total_rows = total
        batch.success_rows = len(success_tasks)
        batch.fail_rows = len(error_rows)
        batch.error_report_path = error_report
        batch.status = "completed" if success_tasks or not error_rows else "completed"

        write_operation_log(
            user_id=user_id,
            action="import_excel",
            target_type="import_batch",
            target_id=batch.id,
            detail={
                "batch_no": batch_no,
                "file_name": safe_name,
                "total_rows": total,
                "success_rows": len(success_tasks),
                "fail_rows": len(error_rows),
            },
        )
        db.session.commit()
        return batch
