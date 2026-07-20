"""分类规则导入与查询服务。"""

from datetime import datetime

from openpyxl import load_workbook

from app.extensions import db
from app.models.rule import (
    ClassificationFieldRule,
    ClassificationRuleChangeLog,
    ClassificationRuleNode,
    ClassificationRuleVersion,
)
from app.services.operation_log_service import write_operation_log
from app.utils.excel_helper import cell_str

# 级联字段顺序（大类～包装方式）
CASCADE_FIELDS = ["大类", "区隔", "类别", "主材质", "辅材质", "包装方式"]

FIELD_TO_NODE_ATTR = {
    "大类": "category_large",
    "区隔": "category_segment",
    "类别": "category_type",
    "主材质": "material_main",
    "辅材质": "material_aux",
    "包装方式": "packaging",
}

FIELD_TO_PATH_ATTR = {
    "大类": "path_large",
    "区隔": "path_segment",
    "类别": "path_type",
    "主材质": "path_material_main",
    "辅材质": "path_material_aux",
    "包装方式": "path_packaging",
    "尺寸": "path_size",
    "卷数": "path_roll",
}

SUPPLEMENT_SHEETS = ["辅材质", "包装方式", "尺寸", "卷数", "总入数"]

# 各补充 Sheet 的路径列与选项列（选项列不得写入路径字段）
SUPPLEMENT_SHEET_CONFIG = {
    "辅材质": {
        "option_col": 5,
        "path_cols": {
            1: "path_large",
            2: "path_segment",
            3: "path_type",
            4: "path_material_main",
        },
    },
    "包装方式": {
        "option_col": 6,
        "path_cols": {
            1: "path_large",
            2: "path_segment",
            3: "path_type",
            4: "path_material_main",
            5: "path_material_aux",
        },
    },
    "尺寸": {
        "option_col": 7,
        "path_cols": {
            1: "path_large",
            2: "path_segment",
            3: "path_type",
            4: "path_material_main",
            5: "path_material_aux",
            6: "path_packaging",
        },
    },
    "卷数": {
        "option_col": 8,
        "path_cols": {
            1: "path_large",
            2: "path_segment",
            3: "path_type",
            4: "path_material_main",
            5: "path_material_aux",
            6: "path_packaging",
            7: "path_size",
        },
    },
    "总入数": {
        "option_col": 9,
        "path_cols": {
            1: "path_large",
            2: "path_segment",
            3: "path_type",
            4: "path_material_main",
            5: "path_material_aux",
            6: "path_packaging",
            7: "path_size",
            8: "path_roll",
        },
    },
}

# 补充规则匹配时，仅使用当前字段之前的路径前缀
FIELD_MATCH_ATTRS = {
    "辅材质": ["path_large", "path_segment", "path_type", "path_material_main"],
    "包装方式": [
        "path_large",
        "path_segment",
        "path_type",
        "path_material_main",
        "path_material_aux",
    ],
    "尺寸": [
        "path_large",
        "path_segment",
        "path_type",
        "path_material_main",
        "path_material_aux",
        "path_packaging",
    ],
    "卷数": [
        "path_large",
        "path_segment",
        "path_type",
        "path_material_main",
        "path_material_aux",
        "path_packaging",
        "path_size",
    ],
    "总入数": [
        "path_large",
        "path_segment",
        "path_type",
        "path_material_main",
        "path_material_aux",
        "path_packaging",
        "path_size",
        "path_roll",
    ],
}

ATTR_TO_PATH_FIELD = {
    "path_large": ("大类", False),
    "path_segment": ("区隔", True),
    "path_type": ("类别", False),
    "path_material_main": ("主材质", False),
    "path_material_aux": ("辅材质", False),
    "path_packaging": ("包装方式", False),
    "path_size": ("尺寸", False),
    "path_roll": ("卷数", False),
}


class RuleService:
    """分类规则业务逻辑。"""

    def _pick_first_unique_hint(self, rules: list) -> str:
        """从匹配规则里提取“去重后的第一个 hint”（优先非空）。

        说明：多个 ClassificationFieldRule 可能同时匹配同一路径，且 hint 可能重复。
        业务要求：去重后取第一个，并且尽量取到非空值。
        """
        seen = set()
        for r in rules:
            hint = (r.hint or "").strip()
            if not hint:
                continue
            if hint in seen:
                continue
            seen.add(hint)
            return hint
        return ""

    def get_latest_version(self) -> ClassificationRuleVersion | None:
        """获取最新规则版本。"""
        return (
            ClassificationRuleVersion.query.order_by(
                ClassificationRuleVersion.id.desc()
            ).first()
        )

    def _normalize_mode(self, value) -> str:
        """统一 0/1 模式标记。"""
        text = cell_str(value)
        if text in ("0", "1"):
            return text
        if text in ("0.0", "1.0"):
            return text.split(".")[0]
        return text

    def _cell_value(self, value):
        """规范化单元格值。"""
        if value is None:
            return None
        text = cell_str(value)
        return text if text else None

    def import_from_excel(self, file_path: str, user_id: int, remark: str = "") -> dict:
        """从规则 Excel 导入并生成新版本。"""
        wb = load_workbook(file_path, read_only=True, data_only=True)

        if "分类规则" not in wb.sheetnames:
            wb.close()
            raise ValueError("缺少 Sheet「分类规则」")

        version_no = datetime.now().strftime("v%Y%m%d%H%M%S")
        version = ClassificationRuleVersion(
            version_no=version_no,
            remark=remark or "Excel 导入",
            created_by=user_id,
        )
        db.session.add(version)
        db.session.flush()

        # 解析主规则树
        ws = wb["分类规则"]
        nodes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            nodes.append(
                ClassificationRuleNode(
                    version_id=version.id,
                    category_large=self._cell_value(row[0]),
                    category_segment=self._cell_value(row[1]),
                    category_type=self._cell_value(row[2]),
                    material_main=self._cell_value(row[3]),
                    material_aux=self._cell_value(row[4]),
                    packaging=self._cell_value(row[5]),
                    size_option=self._cell_value(row[6]),
                    roll_input_mode=self._normalize_mode(row[7]) if row[7] is not None else None,
                    total_input_mode=self._normalize_mode(row[8]) if row[8] is not None else None,
                    is_active=True,
                )
            )
        if nodes:
            db.session.bulk_save_objects(nodes)

        wb.close()
        field_rules = self._parse_supplement_sheets(file_path, version.id)
        if field_rules:
            db.session.bulk_save_objects(field_rules)

        db.session.add(
            ClassificationRuleChangeLog(
                rule_type="version",
                rule_id=version.id,
                action="import",
                before_json=None,
                after_json={
                    "version_no": version_no,
                    "node_count": len(nodes),
                    "field_rule_count": len(field_rules),
                },
                operator_id=user_id,
            )
        )
        write_operation_log(
            user_id=user_id,
            action="import_rules",
            target_type="rule_version",
            target_id=version.id,
            detail={"version_no": version_no, "nodes": len(nodes), "field_rules": len(field_rules)},
        )
        db.session.commit()

        return {
            "version_id": version.id,
            "version_no": version_no,
            "node_count": len(nodes),
            "field_rule_count": len(field_rules),
        }

    def _parse_supplement_sheets(self, file_path: str, version_id: int) -> list:
        """解析补充规则 Sheet。"""
        wb = load_workbook(file_path, read_only=True, data_only=True)
        result = []

        for sheet_name in SUPPLEMENT_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            config = SUPPLEMENT_SHEET_CONFIG[sheet_name]
            option_col = config["option_col"]
            path_cols = config["path_cols"]

            for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
                if not row or all(c is None or cell_str(c) == "" for c in row):
                    continue
                mode = self._normalize_mode(row[0])
                if mode not in ("0", "1"):
                    continue

                rule_kwargs = {
                    "version_id": version_id,
                    "field_name": sheet_name,
                    "input_mode": mode,
                    "option_value": self._cell_value(row[option_col])
                    if len(row) > option_col
                    else None,
                    "hint": self._cell_value(row[10]) if len(row) > 10 else None,
                    "is_active": True,
                }
                for col_idx, attr_name in path_cols.items():
                    rule_kwargs[attr_name] = (
                        self._cell_value(row[col_idx]) if len(row) > col_idx else None
                    )

                result.append(ClassificationFieldRule(**rule_kwargs))

        wb.close()
        return result

    def _filter_nodes(self, version_id: int, path: dict) -> list:
        """按已选路径过滤主规则节点。"""
        nodes = ClassificationRuleNode.query.filter_by(
            version_id=version_id, is_active=True
        ).all()
        matched = []
        segments = path.get("区隔") or []
        if isinstance(segments, str):
            segments = [segments]

        for node in nodes:
            ok = True
            for field in CASCADE_FIELDS:
                attr = FIELD_TO_NODE_ATTR[field]
                selected = path.get(field)
                node_val = getattr(node, attr)
                if selected is None or selected == "" or selected == []:
                    continue
                if field == "区隔":
                    if segments and node_val not in segments:
                        ok = False
                        break
                else:
                    if str(node_val) != str(selected):
                        ok = False
                        break
            if ok:
                matched.append(node)
        return matched

    def _path_specificity(self, rule: ClassificationFieldRule) -> int:
        """路径前缀匹配优先级（填写列越多越具体）。"""
        score = 0
        for attr in FIELD_TO_PATH_ATTR.values():
            if getattr(rule, attr):
                score += 1
        return score

    def _match_field_rule(self, rule: ClassificationFieldRule, path: dict) -> bool:
        """判断补充规则是否匹配当前路径（仅匹配该字段之前的路径前缀）。"""
        segments = path.get("区隔") or []
        if isinstance(segments, str):
            segments = [segments]

        attrs = FIELD_MATCH_ATTRS.get(rule.field_name, list(ATTR_TO_PATH_FIELD.keys()))
        for attr in attrs:
            field, is_segment = ATTR_TO_PATH_FIELD[attr]
            rule_val = getattr(rule, attr)
            if not rule_val:
                continue
            path_val = path.get(field)
            if is_segment:
                if not segments or str(rule_val) not in [str(s) for s in segments]:
                    return False
            elif path_val is None or str(path_val) != str(rule_val):
                return False
        return True

    def _find_field_rule(self, field_name: str, path: dict, version_id: int):
        """查找最匹配的补充规则。"""
        matched = self._matched_field_rules(field_name, path, version_id)
        return matched[0] if matched else None

    def _matched_field_rules(
        self, field_name: str, path: dict, version_id: int
    ) -> list:
        """获取匹配路径的补充规则（按路径具体程度降序）。"""
        rules = ClassificationFieldRule.query.filter_by(
            version_id=version_id, field_name=field_name, is_active=True
        ).all()
        matched = [r for r in rules if self._match_field_rule(r, path)]
        matched.sort(key=self._path_specificity, reverse=True)
        return matched

    def get_cascade_options(self, field: str, path: dict) -> dict:
        """获取级联下拉选项（大类～包装方式）。"""
        version = self.get_latest_version()
        if not version:
            return {"field": field, "options": [], "input_mode": "select", "hint": ""}

        if field not in CASCADE_FIELDS:
            return {"field": field, "options": [], "input_mode": "select", "hint": ""}

        # 辅材质/包装方式：可能为文本框或下拉，提示词均来自补充规则 hint
        field_rule_hint = ""
        if field in ("辅材质", "包装方式"):
            matched = self._matched_field_rules(field, path, version.id)
            text_rules = [r for r in matched if r.input_mode == "0"]
            select_rules = [r for r in matched if r.input_mode == "1"]
            if text_rules:
                return {
                    "field": field,
                    "options": [],
                    "input_mode": "text",
                    "hint": self._pick_first_unique_hint(text_rules),
                }
            field_rule_hint = self._pick_first_unique_hint(select_rules)

        nodes = self._filter_nodes(version.id, path)
        attr = FIELD_TO_NODE_ATTR[field]
        options = []
        seen = set()
        for node in nodes:
            val = getattr(node, attr)
            if val and val not in seen:
                seen.add(val)
                options.append(val)

        return {
            "field": field,
            "options": options,
            "input_mode": "select",
            "hint": field_rule_hint,
        }

    def get_field_meta(self, field: str, path: dict) -> dict:
        """获取尺寸/卷数/总入数（及文本模式字段）的控件元数据。"""
        version = self.get_latest_version()
        if not version:
            return {"field": field, "input_mode": "text", "options": [], "hint": ""}

        nodes = self._filter_nodes(version.id, path)

        if field == "尺寸":
            size_values = [
                n.size_option
                for n in nodes
                if n.size_option and str(n.size_option) not in ("0", "1")
            ]
            if size_values:
                matched = self._matched_field_rules("尺寸", path, version.id)
                return {
                    "field": field,
                    "input_mode": "select",
                    "options": list(dict.fromkeys(size_values)),
                    "hint": self._pick_first_unique_hint(matched),
                    "multiple": False,
                }
            # 主表尺寸列为 0/1，走「尺寸」Sheet
            meta = self._meta_from_field_rules("尺寸", path, version.id, default_mode="text")
            if meta.get("input_mode") == "select":
                meta["multiple"] = True
            else:
                meta["multiple"] = False
            return meta

        if field == "卷数":
            modes = {n.roll_input_mode for n in nodes if n.roll_input_mode in ("0", "1")}
            mode = modes.pop() if len(modes) == 1 else (modes.pop() if modes else "0")
            if mode == "1":
                meta = self._meta_from_field_rules("卷数", path, version.id, default_mode="select")
                meta["multiple"] = False
                return meta
            matched = self._matched_field_rules("卷数", path, version.id)
            return {
                "field": field,
                "input_mode": "text",
                "options": [],
                "hint": self._pick_first_unique_hint(matched),
            }

        if field == "总入数":
            modes = {n.total_input_mode for n in nodes if n.total_input_mode in ("0", "1")}
            mode = modes.pop() if len(modes) == 1 else (modes.pop() if modes else "0")
            if mode == "1":
                meta = self._meta_from_field_rules("总入数", path, version.id, default_mode="select")
                meta["multiple"] = False
                return meta
            matched = self._matched_field_rules("总入数", path, version.id)
            return {
                "field": field,
                "input_mode": "text",
                "options": [],
                "hint": self._pick_first_unique_hint(matched),
            }

        return {"field": field, "input_mode": "text", "options": [], "hint": ""}

    def _meta_from_field_rules(
        self, field_name: str, path: dict, version_id: int, default_mode: str
    ) -> dict:
        """从补充 Sheet 规则汇总控件信息。"""
        matched = self._matched_field_rules(field_name, path, version_id)
        if not matched:
            return {"field": field_name, "input_mode": default_mode, "options": [], "hint": ""}

        # 取下拉选项（mode=1 的所有匹配行）
        select_rules = [r for r in matched if r.input_mode == "1" and r.option_value]
        text_rules = [r for r in matched if r.input_mode == "0"]
        if select_rules:
            options = list(dict.fromkeys(r.option_value for r in select_rules if r.option_value))
            hint = self._pick_first_unique_hint(select_rules)
            return {
                "field": field_name,
                "input_mode": "select",
                "options": options,
                "hint": hint,
                "multiple": field_name == "尺寸",
            }

        text_rules.sort(key=self._path_specificity, reverse=True)
        return {
            "field": field_name,
            "input_mode": "text",
            "options": [],
            "hint": self._pick_first_unique_hint(text_rules),
            "multiple": False,
        }
